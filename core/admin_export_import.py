# ==================================================
# FILE: core/admin_export_import.py (VERSI DIPERBAIKI - BYTES)
# ==================================================

import csv
import json
import openpyxl
import os
import zipfile
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.db import connection, transaction
from io import TextIOWrapper, BytesIO, StringIO

# ===== IMPORT MODEL DENGAN BENAR =====
from django.contrib.auth.models import User
from core.models import (
    KodeBarang, Supplier, StockItem, StockPurchase, StockPurchaseItem,
    MenuCategory, Menu, MenuIngredient, Order, OrderItem, Profile
)


# ==================== EXPORT ALL DATA ====================

@staff_member_required
def export_all_data(request):
    """
    Export SEMUA data ke dalam satu file ZIP berisi CSV/Excel
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Buat ZIP file dalam memory
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # 1. Export KodeBarang
        csv_data = export_kode_barang_to_csv()
        zip_file.writestr('01_kode_barang.csv', csv_data)
        
        # 2. Export Supplier
        csv_data = export_supplier_to_csv()
        zip_file.writestr('02_supplier.csv', csv_data)
        
        # 3. Export StockItem
        csv_data = export_stock_item_to_csv()
        zip_file.writestr('03_stock_item.csv', csv_data)
        
        # 4. Export StockPurchase
        csv_data = export_stock_purchase_to_csv()
        zip_file.writestr('04_stock_purchase.csv', csv_data)
        
        # 5. Export StockPurchaseItem
        csv_data = export_stock_purchase_item_to_csv()
        zip_file.writestr('05_stock_purchase_item.csv', csv_data)
        
        # 6. Export MenuCategory
        csv_data = export_menu_category_to_csv()
        zip_file.writestr('06_menu_category.csv', csv_data)
        
        # 7. Export Menu
        csv_data = export_menu_to_csv()
        zip_file.writestr('07_menu.csv', csv_data)
        
        # 8. Export MenuIngredient
        csv_data = export_menu_ingredient_to_csv()
        zip_file.writestr('08_menu_ingredient.csv', csv_data)
        
        # 9. Export Order
        csv_data = export_order_to_csv()
        zip_file.writestr('09_order.csv', csv_data)
        
        # 10. Export OrderItem
        csv_data = export_order_item_to_csv()
        zip_file.writestr('10_order_item.csv', csv_data)
        
        # 11. Export User & Profile
        csv_data = export_user_to_csv()
        zip_file.writestr('11_user.csv', csv_data)
        
        # 12. Buat file INFO
        info = f"""MBAYAR POS - BACKUP DATABASE
Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}
User: {request.user.username}

File-file dalam backup ini:
- 01_kode_barang.csv : Master kode barang
- 02_supplier.csv : Data supplier
- 03_stock_item.csv : Item stok
- 04_stock_purchase.csv : Pembelian stok (header)
- 05_stock_purchase_item.csv : Detail pembelian
- 06_menu_category.csv : Kategori menu
- 07_menu.csv : Data menu
- 08_menu_ingredient.csv : Bahan-bahan menu
- 09_order.csv : Transaksi penjualan
- 10_order_item.csv : Detail transaksi
- 11_user.csv : Data user

Cara restore:
1. Hapus semua data yang ada (opsional)
2. Import file-file CSV ini berurutan dari 01 sampai 11
"""
        zip_file.writestr('README_BACKUP.txt', info.encode('utf-8'))
    
    # Siapkan response
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="mbayar_backup_{timestamp}.zip"'
    
    return response


# ==================== FUNGSI EXPORT KE CSV (RETURN BYTES) ====================

def export_kode_barang_to_csv():
    """Export KodeBarang ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'KODE', 'NAMA', 'KETERANGAN', 'CREATED_AT'])
    
    for obj in KodeBarang.objects.all():
        writer.writerow([
            obj.id,
            obj.kode,
            obj.nama,
            obj.keterangan,
            obj.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return output.getvalue().encode('utf-8')


def export_supplier_to_csv():
    """Export Supplier ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'NAMA', 'KONTAK', 'TELEPON', 'ALAMAT', 'CREATED_AT'])
    
    for obj in Supplier.objects.all():
        writer.writerow([
            obj.id,
            obj.name,
            obj.contact,
            obj.phone,
            obj.address,
            obj.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return output.getvalue().encode('utf-8')


def export_stock_item_to_csv():
    """Export StockItem ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'KODE_BARANG_ID', 'NAMA', 'UNIT', 'STOK', 'MIN_STOK',
        'SUPPLIER_ID', 'MARKUP_PERSEN', 'MARKUP_NOMINAL',
        'HARGA_BELI_TERAKHIR', 'HARGA_JUAL', 'HARGA_GOFOOD'
    ])
    
    for obj in StockItem.objects.all():
        writer.writerow([
            obj.id,
            obj.kode_barang_id,
            obj.name,
            obj.unit,
            obj.stock,
            obj.min_stock,
            obj.supplier_id or '',
            obj.markup_persen,
            obj.markup_nominal,
            obj.harga_beli_terakhir,
            obj.harga_jual,
            obj.harga_gofood
        ])
    
    return output.getvalue().encode('utf-8')


def export_stock_purchase_to_csv():
    """Export StockPurchase ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'INVOICE_NO', 'SUPPLIER_ID', 'DATE', 'TOTAL_AMOUNT', 'NOTES', 'CREATED_BY_ID'])
    
    for obj in StockPurchase.objects.all():
        writer.writerow([
            obj.id,
            obj.invoice_no,
            obj.supplier_id or '',
            obj.date.strftime('%Y-%m-%d %H:%M:%S'),
            obj.total_amount,
            obj.notes,
            obj.created_by_id or ''
        ])
    
    return output.getvalue().encode('utf-8')


def export_stock_purchase_item_to_csv():
    """Export StockPurchaseItem ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'PURCHASE_ID', 'KODE_BARANG_ID', 'JUMLAH', 'HARGA_TOTAL',
        'CARA_HITUNG', 'HARGA_UNIT', 'SUBTOTAL'
    ])
    
    for obj in StockPurchaseItem.objects.all():
        writer.writerow([
            obj.id,
            obj.purchase_id,
            obj.kode_barang_id,
            obj.jumlah,
            obj.harga_total,
            obj.cara_hitung,
            obj.harga_unit,
            obj.subtotal
        ])
    
    return output.getvalue().encode('utf-8')


def export_menu_category_to_csv():
    """Export MenuCategory ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'NAMA', 'DESKRIPSI'])
    
    for obj in MenuCategory.objects.all():
        writer.writerow([obj.id, obj.name, obj.description])
    
    return output.getvalue().encode('utf-8')


def export_menu_to_csv():
    """Export Menu ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'CODE', 'NAME', 'CATEGORY_ID', 'IS_AVAILABLE',
        'MARKUP_PERSEN', 'MARKUP_NOMINAL', 'TOTAL_MODAL', 'SELLING_PRICE'
    ])
    
    for obj in Menu.objects.all():
        writer.writerow([
            obj.id,
            obj.code,
            obj.name,
            obj.category_id or '',
            1 if obj.is_available else 0,
            obj.markup_persen,
            obj.markup_nominal,
            obj.total_modal,
            obj.selling_price
        ])
    
    return output.getvalue().encode('utf-8')


def export_menu_ingredient_to_csv():
    """Export MenuIngredient ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'MENU_ID', 'STOCK_ITEM_ID', 'QUANTITY_USED', 'NOTES'])
    
    for obj in MenuIngredient.objects.all():
        writer.writerow([
            obj.id,
            obj.menu_id,
            obj.stock_item_id,
            obj.quantity_used,
            obj.notes
        ])
    
    return output.getvalue().encode('utf-8')


def export_order_to_csv():
    """Export Order ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'ORDER_NO', 'CASHIER_ID', 'CUSTOMER_NAME', 'ORDER_DATE',
        'STATUS', 'PAYMENT_METHOD', 'SUBTOTAL', 'TAX', 'DISCOUNT',
        'TOTAL', 'AMOUNT_PAID', 'CHANGE'
    ])
    
    for obj in Order.objects.all():
        writer.writerow([
            obj.id,
            obj.order_no,
            obj.cashier_id or '',
            obj.customer_name,
            obj.order_date.strftime('%Y-%m-%d %H:%M:%S'),
            obj.status,
            obj.payment_method,
            obj.subtotal,
            obj.tax,
            obj.discount,
            obj.total,
            obj.amount_paid,
            obj.change
        ])
    
    return output.getvalue().encode('utf-8')


def export_order_item_to_csv():
    """Export OrderItem ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'ORDER_ID', 'MENU_ID', 'MENU_NAME', 'QUANTITY', 'PRICE', 'SUBTOTAL', 'NOTES'])
    
    for obj in OrderItem.objects.all():
        writer.writerow([
            obj.id,
            obj.order_id,
            obj.menu_id or '',
            obj.menu_name,
            obj.quantity,
            obj.price,
            obj.subtotal,
            obj.notes
        ])
    
    return output.getvalue().encode('utf-8')


def export_user_to_csv():
    """Export User dan Profile ke CSV (return bytes)"""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'USERNAME', 'FIRST_NAME', 'LAST_NAME', 'EMAIL', 'PASSWORD',
        'IS_STAFF', 'IS_ACTIVE', 'DATE_JOINED', 'ROLE', 'PHONE'
    ])
    
    for user in User.objects.all():
        try:
            profile = Profile.objects.get(user=user)
            role = profile.role
            phone = profile.phone
        except Profile.DoesNotExist:
            role = 'user'
            phone = ''
        
        writer.writerow([
            user.id,
            user.username,
            user.first_name,
            user.last_name,
            user.email,
            user.password,
            1 if user.is_staff else 0,
            1 if user.is_active else 0,
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            role,
            phone
        ])
    
    return output.getvalue().encode('utf-8')


# ==================== IMPORT DATA ====================

@staff_member_required
def import_data_page(request):
    """Halaman untuk import data"""
    return render(request, 'admin/import_data.html')


@staff_member_required
def import_from_zip(request):
    """
    Import data dari file ZIP backup
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method tidak diizinkan'}, status=405)
    
    try:
        zip_file = request.FILES.get('zip_file')
        delete_existing = request.POST.get('delete_existing') == 'on'
        
        if not zip_file:
            return JsonResponse({'error': 'File ZIP harus diisi'}, status=400)
        
        # Baca ZIP file
        zip_buffer = BytesIO(zip_file.read())
        results = {
            'success': True,
            'deleted': False,
            'imported': {},
            'errors': []
        }
        
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            
            # ===== HAPUS DATA LAMA JIKA DIPILIH =====
            if delete_existing:
                with transaction.atomic():
                    delete_all_data()
                    results['deleted'] = True
            
            # ===== IMPORT DATA BERURUTAN =====
            file_order = [
                '01_kode_barang.csv',
                '02_supplier.csv',
                '03_stock_item.csv',
                '04_stock_purchase.csv',
                '05_stock_purchase_item.csv',
                '06_menu_category.csv',
                '07_menu.csv',
                '08_menu_ingredient.csv',
                '09_order.csv',
                '10_order_item.csv',
                '11_user.csv'
            ]
            
            for filename in file_order:
                try:
                    if filename in zf.namelist():
                        with zf.open(filename) as f:
                            # Baca file sebagai text
                            text_file = TextIOWrapper(f, encoding='utf-8')
                            count = import_csv_file(text_file, filename)
                            results['imported'][filename] = count
                    else:
                        results['imported'][filename] = 0
                        results['errors'].append(f'File {filename} tidak ditemukan dalam ZIP')
                except Exception as e:
                    results['errors'].append(f'Error import {filename}: {str(e)}')
            
            # ===== HITUNG ULANG HARGA JUAL =====
            recalculate_all_prices()
        
        return JsonResponse(results)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def delete_all_data():
    """Hapus semua data dalam urutan yang benar"""
    print("🗑️ Menghapus semua data...")
    
    # Hapus child tables dulu
    OrderItem.objects.all().delete()
    Order.objects.all().delete()
    MenuIngredient.objects.all().delete()
    Menu.objects.all().delete()
    MenuCategory.objects.all().delete()
    StockPurchaseItem.objects.all().delete()
    StockPurchase.objects.all().delete()
    StockItem.objects.all().delete()
    Supplier.objects.all().delete()
    KodeBarang.objects.all().delete()
    
    # Hapus profile dulu, baru user
    Profile.objects.all().delete()
    User.objects.filter(is_superuser=False).delete()
    
    print("✅ Semua data terhapus")


def import_csv_file(file_obj, filename):
    """Import satu file CSV"""
    reader = csv.reader(file_obj)
    next(reader)  # Skip header
    
    count = 0
    
    if filename == '01_kode_barang.csv':
        count = import_kode_barang(reader)
    elif filename == '02_supplier.csv':
        count = import_supplier(reader)
    elif filename == '03_stock_item.csv':
        count = import_stock_item(reader)
    elif filename == '04_stock_purchase.csv':
        count = import_stock_purchase(reader)
    elif filename == '05_stock_purchase_item.csv':
        count = import_stock_purchase_item(reader)
    elif filename == '06_menu_category.csv':
        count = import_menu_category(reader)
    elif filename == '07_menu.csv':
        count = import_menu(reader)
    elif filename == '08_menu_ingredient.csv':
        count = import_menu_ingredient(reader)
    elif filename == '09_order.csv':
        count = import_order(reader)
    elif filename == '10_order_item.csv':
        count = import_order_item(reader)
    elif filename == '11_user.csv':
        count = import_user(reader)
    
    return count


def import_kode_barang(reader):
    """Import data KodeBarang"""
    count = 0
    for row in reader:
        if len(row) >= 5:
            KodeBarang.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'kode': row[1],
                    'nama': row[2],
                    'keterangan': row[3],
                    'created_at': row[4]
                }
            )
            count += 1
    return count


def import_supplier(reader):
    """Import data Supplier"""
    count = 0
    for row in reader:
        if len(row) >= 6:
            Supplier.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'name': row[1],
                    'contact': row[2],
                    'phone': row[3],
                    'address': row[4],
                    'created_at': row[5]
                }
            )
            count += 1
    return count


def import_stock_item(reader):
    """Import data StockItem"""
    count = 0
    for row in reader:
        if len(row) >= 12:
            supplier_id = int(row[6]) if row[6] else None
            StockItem.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'kode_barang_id': int(row[1]),
                    'name': row[2],
                    'unit': row[3],
                    'stock': float(row[4]),
                    'min_stock': float(row[5]),
                    'supplier_id': supplier_id,
                    'markup_persen': row[7],
                    'markup_nominal': row[8],
                    'harga_beli_terakhir': row[9],
                    'harga_jual': row[10],
                    'harga_gofood': row[11]
                }
            )
            count += 1
    return count


def import_stock_purchase(reader):
    """Import data StockPurchase"""
    count = 0
    for row in reader:
        if len(row) >= 7:
            supplier_id = int(row[2]) if row[2] else None
            created_by_id = int(row[6]) if row[6] else None
            StockPurchase.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'invoice_no': row[1],
                    'supplier_id': supplier_id,
                    'date': row[3],
                    'total_amount': row[4],
                    'notes': row[5],
                    'created_by_id': created_by_id
                }
            )
            count += 1
    return count


def import_stock_purchase_item(reader):
    """Import data StockPurchaseItem"""
    count = 0
    for row in reader:
        if len(row) >= 8:
            StockPurchaseItem.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'purchase_id': int(row[1]),
                    'kode_barang_id': int(row[2]),
                    'jumlah': float(row[3]),
                    'harga_total': row[4],
                    'cara_hitung': row[5],
                    'harga_unit': row[6],
                    'subtotal': row[7]
                }
            )
            count += 1
    return count


def import_menu_category(reader):
    """Import data MenuCategory"""
    count = 0
    for row in reader:
        if len(row) >= 3:
            MenuCategory.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'name': row[1],
                    'description': row[2]
                }
            )
            count += 1
    return count


def import_menu(reader):
    """Import data Menu"""
    count = 0
    for row in reader:
        if len(row) >= 9:
            category_id = int(row[3]) if row[3] else None
            Menu.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'code': row[1],
                    'name': row[2],
                    'category_id': category_id,
                    'is_available': bool(int(row[4])),
                    'markup_persen': row[5],
                    'markup_nominal': row[6],
                    'total_modal': row[7],
                    'selling_price': row[8]
                }
            )
            count += 1
    return count


def import_menu_ingredient(reader):
    """Import data MenuIngredient"""
    count = 0
    for row in reader:
        if len(row) >= 5:
            MenuIngredient.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'menu_id': int(row[1]),
                    'stock_item_id': int(row[2]),
                    'quantity_used': float(row[3]),
                    'notes': row[4]
                }
            )
            count += 1
    return count


def import_order(reader):
    """Import data Order"""
    count = 0
    for row in reader:
        if len(row) >= 13:
            cashier_id = int(row[2]) if row[2] else None
            Order.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'order_no': row[1],
                    'cashier_id': cashier_id,
                    'customer_name': row[3],
                    'order_date': row[4],
                    'status': row[5],
                    'payment_method': row[6],
                    'subtotal': row[7],
                    'tax': row[8],
                    'discount': row[9],
                    'total': row[10],
                    'amount_paid': row[11],
                    'change': row[12]
                }
            )
            count += 1
    return count


def import_order_item(reader):
    """Import data OrderItem"""
    count = 0
    for row in reader:
        if len(row) >= 8:
            menu_id = int(row[2]) if row[2] else None
            OrderItem.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'order_id': int(row[1]),
                    'menu_id': menu_id,
                    'menu_name': row[3],
                    'quantity': int(row[4]),
                    'price': row[5],
                    'subtotal': row[6],
                    'notes': row[7]
                }
            )
            count += 1
    return count


def import_user(reader):
    """Import data User dan Profile"""
    count = 0
    for row in reader:
        if len(row) >= 11:
            user, created = User.objects.update_or_create(
                id=int(row[0]),
                defaults={
                    'username': row[1],
                    'first_name': row[2],
                    'last_name': row[3],
                    'email': row[4],
                    'password': row[5],
                    'is_staff': bool(int(row[6])),
                    'is_active': bool(int(row[7])),
                    'date_joined': row[8]
                }
            )
            
            # Update profile
            Profile.objects.update_or_create(
                user=user,
                defaults={
                    'role': row[9],
                    'phone': row[10]
                }
            )
            count += 1
    return count


def recalculate_all_prices():
    """Hitung ulang semua harga setelah import"""
    print("🧮 Menghitung ulang semua harga...")
    
    # Update harga jual semua stock item
    for item in StockItem.objects.all():
        item.hitung_harga_jual()
        item.save()
    
    # Update harga jual semua menu
    for menu in Menu.objects.all():
        menu.calculate_prices()
        menu.save()
    
    print("✅ Perhitungan ulang selesai")


# ==================== TEMPLATE DOWNLOAD ====================

@staff_member_required
def download_import_template(request, template_type):
    """Download template Excel untuk import"""
    
    if template_type == 'kode_barang':
        filename = "template_kode_barang.xlsx"
        headers = ['KODE', 'NAMA', 'KETERANGAN']
        sample_data = [
            ['B001', 'Gula Pasir', 'Gula putih 1kg'],
            ['B002', 'Telur', 'Telur ayam negeri'],
            ['B003', 'Minyak Goreng', 'Minyak sawit 1L']
        ]
    elif template_type == 'supplier':
        filename = "template_supplier.xlsx"
        headers = ['NAMA', 'KONTAK', 'TELEPON', 'ALAMAT']
        sample_data = [
            ['PT Sumber Makmur', 'Budi', '08123456789', 'Jl. Sudirman No. 123'],
            ['CV Berkah Jaya', 'Ani', '08129876543', 'Jl. Gatot Subroto No. 45']
        ]
    elif template_type == 'stock_item':
        filename = "template_stok.xlsx"
        headers = ['KODE_BARANG', 'NAMA', 'UNIT', 'STOK_AWAL', 'MIN_STOK', 'MARKUP_PERSEN']
        sample_data = [
            ['B001', 'Gula Pasir', 'kg', 10, 2, 30],
            ['B002', 'Telur', 'pcs', 50, 10, 25]
        ]
    elif template_type == 'menu':
        filename = "template_menu.xlsx"
        headers = ['KODE_MENU', 'NAMA_MENU', 'KATEGORI', 'KODE_BAHAN', 'JUMLAH', 'CATATAN']
        sample_data = [
            ['M001', 'Nasi Goreng', 'Makanan', 'B001', 0.25, 'kg'],
            ['M001', 'Nasi Goreng', 'Makanan', 'B002', 2, 'butir'],
            ['M002', 'Es Teh', 'Minuman', 'B004', 1, 'teh celup']
        ]
    else:
        return HttpResponse('Tipe template tidak dikenal', status=400)
    
    # Buat Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="8B1E2D", end_color="8B1E2D", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Sample data
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


# ==================== CLEAR DATABASE ====================

@staff_member_required
def clear_database(request):
    """Halaman untuk menghapus semua data"""
    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        
        if confirm == 'HAPUS SEMUA':
            delete_all_data()
            messages.success(request, '✅ Semua database telah dihapus!')
        else:
            messages.error(request, '❌ Konfirmasi tidak sesuai. Ketik "HAPUS SEMUA" untuk konfirmasi')
        
        return redirect('admin_data_management')
    
    return render(request, 'admin/clear_database.html')