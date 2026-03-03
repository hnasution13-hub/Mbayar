# ==================================================
# FILE: Mbayar/core/utils/excel_exporters.py
# PATH: D:/Project Pyton/Mbayar/core/utils/excel_exporters.py
# FUNGSI: Class untuk export ke Excel
# ==================================================

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

class ExcelExporter:
    """Base class untuk export Excel"""
    
    def __init__(self, sheet_title="Laporan"):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_title
        self.current_row = 1
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup style yang akan dipakai"""
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="8B1E2D", end_color="8B1E2D", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def add_header(self, headers):
        """Tambah header row"""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        self.current_row += 1
    
    def add_row(self, data, formats=None):
        """Tambah data row"""
        for col, value in enumerate(data, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=value)
            cell.border = self.border
            
            # Format number jika diperlukan
            if formats and col in formats:
                if formats[col] == 'number':
                    cell.number_format = '#,##0'
                elif formats[col] == 'percentage':
                    cell.number_format = '0.00"%"'
            
            # Alignment
            if isinstance(value, (int, float)) and col > 1:
                cell.alignment = Alignment(horizontal="right")
        
        self.current_row += 1
    
    def add_title(self, title, col_span=5):
        """Tambah judul laporan"""
        cell = self.ws.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        self.ws.merge_cells(
            start_row=self.current_row, start_column=1,
            end_row=self.current_row, end_column=col_span
        )
        cell.alignment = Alignment(horizontal="center")
        self.current_row += 2
    
    def add_period_info(self, start_date, end_date):
        """Tambah info periode"""
        self.ws.cell(row=self.current_row, column=1, 
                    value=f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
        self.current_row += 1
    
    def add_total_row(self, label, value, col_span=2):
        """Tambah row total"""
        label_cell = self.ws.cell(row=self.current_row, column=col_span, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        
        value_cell = self.ws.cell(row=self.current_row, column=col_span + 1, value=value)
        value_cell.font = Font(bold=True)
        value_cell.number_format = '#,##0'
        value_cell.border = self.border
        self.current_row += 1
    
    def add_footer(self):
        """Tambah footer (timestamp)"""
        self.current_row += 1
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')
        self.ws.cell(row=self.current_row, column=1, 
                    value=f"Diekspor pada: {timestamp}")
    
    def auto_width(self):
        """Auto-adjust column width"""
        for col in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_response(self, filename):
        """Return HttpResponse dengan file Excel"""
        self.auto_width()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        self.wb.save(response)
        return response


class SalesExcelExporter(ExcelExporter):
    """Export laporan penjualan"""
    
    def __init__(self):
        super().__init__("Laporan Penjualan")
    
    def export(self, orders, start_date, end_date):
        self.add_title("LAPORAN PENJUALAN")
        self.add_period_info(start_date, end_date)
        self.current_row += 1
        
        # Header
        headers = ['No', 'Tanggal', 'No. Order', 'Pelanggan', 'Kasir', 'Subtotal', 'Diskon', 'Pajak', 'Total', 'Metode']
        self.add_header(headers)
        
        # Data
        total_penjualan = 0
        for i, order in enumerate(orders, 1):
            subtotal = float(order.subtotal) if order.subtotal else 0
            discount = float(order.discount) if order.discount else 0
            tax = float(order.tax) if order.tax else 0
            total = float(order.total) if order.total else 0
            total_penjualan += total
            
            data = [
                i,
                order.order_date.strftime('%d/%m/%Y %H:%M'),
                order.order_no,
                order.customer_name,
                order.cashier.username if order.cashier else '-',
                subtotal,
                discount,
                tax,
                total,
                order.get_payment_method_display()
            ]
            formats = {6: 'number', 7: 'number', 8: 'number', 9: 'number'}
            self.add_row(data, formats)
        
        # Total
        self.current_row += 1
        self.add_total_row("TOTAL PENJUALAN:", total_penjualan, 8)
        
        # Info tambahan
        self.current_row += 1
        self.ws.cell(row=self.current_row, column=1, value=f"Total Transaksi: {orders.count()}")
        
        self.add_footer()
        return self


class ProfitExcelExporter(ExcelExporter):
    """Export laporan laba rugi"""
    
    def __init__(self):
        super().__init__("Laporan Laba Rugi")
    
    def export(self, detail_items, start_date, end_date, summary):
        self.add_title("LAPORAN LABA RUGI")
        self.add_period_info(start_date, end_date)
        self.current_row += 1
        
        # Header
        headers = ['No', 'Tanggal', 'No. Order', 'Menu', 'Qty', 'Harga Jual', 'HPP', 'Laba', 'Margin %']
        self.add_header(headers)
        
        # Data
        for i, item in enumerate(detail_items[:100], 1):  # Batasi 100 untuk performance
            data = [
                i,
                item['order_date'].strftime('%d/%m/%Y'),
                item['order_no'],
                item['menu_name'],
                item['quantity'],
                item['price'],
                item['hpp'],
                item['laba'],
                item['margin']
            ]
            formats = {6: 'number', 7: 'number', 8: 'number', 9: 'percentage'}
            self.add_row(data, formats)
        
        # Summary
        self.current_row += 2
        self.add_total_row("Total Penjualan:", summary['total_penjualan'], 5)
        self.add_total_row("Total HPP:", summary['total_hpp'], 5)
        self.add_total_row("Laba Kotor:", summary['laba_kotor'], 5)
        
        self.current_row += 1
        self.ws.cell(row=self.current_row, column=6, value=f"Margin: {summary['margin']:.1f}%")
        
        self.add_footer()
        return self